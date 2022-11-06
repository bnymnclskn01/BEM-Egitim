# -*- coding: utf-8 -*-
"""
Created on Sun Nov  6 11:18:47 2022

@author: PC
"""

import openpyxl 

wb = openpyxl.load_workbook("C:\\Users\\ever_\\OneDrive\\Belgeler\\GitHub\\BEM-Egitim\\PYH\\06.11.2022\\İsletme.xlsx")
#wb = openpyxl.load_workbook("İsletme.xlsx")
sheet = wb.sheetnames
Sheet1=wb[sheet[0]]

"""
print(type(sheet[0]))

print(type(Sheet1))

A=Sheet1.cell(1,1).value
print(A)

for row in Sheet1.iter_cols():
    print(row)
    break
"""

column=0
for col in Sheet1.iter_cols():
    column=+1+column
    #print(col)

row=0    
for rw in Sheet1.iter_rows():
    row=1+row
    #print(rw)

Person_Count=row-1
Sucsess_Percent=0
Sucsess_Count=0
UnSucsess_Count=0

for j in range(2,row+1):
   if(Sheet1.cell(j,2).value=="Başarılı"):
       Sucsess_Count=1+Sucsess_Count
   elif(Sheet1.cell(j,2).value=="Başarısız"):
       UnSucsess_Count=1+UnSucsess_Count
       
Sucsess_Percent=(Sucsess_Count/(Sucsess_Count+UnSucsess_Count))*100

print("Kac kisi geldi: {}\nKac islem basarılı oldu: {}\nKac islem basarısız oldu: {}".format(Person_Count, Sucsess_Count,UnSucsess_Count))
print("Kaç işlem yapıldı: {}\nBaşarılı ve Başarısız Yüzde: {}".format(Person_Count, Sucsess_Percent))




