# -*- coding: utf-8 -*-
"""
Created on Sat Nov  5 14:48:53 2022

@author: ever_
"""
"""
dosya=open("dosya.txt","r",encoding=("utf-8"))
oku=dosya.read()
print(oku)

dosya=open("dosya.txt","r",encoding=("utf-8"))
oku=dosya.read(10) #sadece ilk 10 karekteri okur
print(oku)

dosya=open("dosya.txt","r",encoding=("utf-8"))
print(dosya.readline())
print(dosya.readline())

with open("dosya.txt","r",encoding=("utf-8")) as dosya:
    for satir in dosya:
        print(satir)
dosya=open("dosya.txt","r",encoding=("utf-8"))
dizi=dosya.readlines()
print(dizi)

dosya=open("dosya.txt","a",encoding=("utf-8"))
dosya.write("Kerteriz ile Python Dosya İşlemleri\n")
dosya.write("Burasi satır oluyor bilginiz olsun")

with open("dosya.txt","r",encoding=("utf-8")) as dosya:
    for satir in dosya:
        print(satir)

with open("dosya.txt","r",encoding=("utf-8")) as dosya:
    print(dosya.tell())
with open("dosya.txt","r",encoding=("utf-8")) as dosya:
    print(dosya.seek(29))
with open("dosya.txt","r",encoding=("utf-8")) as dosya:
    print(dosya.tell())

with open("dosya.txt","r+",encoding=("utf-8")) as dosya:
    icerik=dosya.read()
    yeniicerik="Bu metini ilk satıra atamak istiyorum.\n"+icerik
    dosya.seek(0)
    dosya.write(yeniicerik)

with open("dosya.txt","r",encoding=("utf-8")) as dosya:
    for satir in dosya:
        print(satir)

dosya=open("dosya.txt","r+",encoding=("utf-8"))
deger=0
for satir in dosya:
    deger=deger+1
print(deger)
with open("dosya.txt","r",encoding=("utf-8")) as f:
    print(f.read().count("\n"))


if os.path.exists("dosya.txt"):
    os.remove("dosya.txt")
else:
    print("Dosya Bulunamadı")

import os
os.rmdir("pyh")

fo=open("dosya.txt","w")
print("Dosya : ",fo.name)
print("Kaplı mı değil mi ? ",fo.closed)
print("Erişim Modu",fo.mode)
from openpyxl import workbook, load_workbook
wb=load_workbook("Dosya.xlsx")
ws=wb.active
ws=wb.sheetnames
print(ws)
ws=wb["Sayfa1"]
print(ws)
print(ws["A1"].value)
print(ws.cell(1, 1).value)
for satir in range(2,5):
    for sutun in range(1,4):
        print("|"+str(ws.cell(satir,sutun).value)+"|",end="")
for satir in range (1,ws.max_row+1):
    for sutun in range(1,ws.max_column+1):
        print("|"+str(ws.cell(satir,sutun).value)+"|",end="")
"""

"""
wb = Workbook()
ws=wb.active
ws.Title="İlk Çalışma Alanı"
ws= wb.create_sheet("Posta Kodları")
ws= wb.create_sheet("Ülkeler")
print(wb.sheetnames)
wb.save("DosyaAdi.xlsx")
"""
"""
from openpyxl import Workbook, load_workbook
wb =load_workbook("DosyaAdi.xlsx")
ws = wb.active
toplam = 0
sayi=0
for i in range(2,755464):
    if ( ws["C"+str(i)].value == None ):
        break
    else:
        print(ws["C"+str(i)].value)
        toplam += int( ws["C"+str(i)].value)
        sayi+=1
        
        
ort=toplam/sayi
print(ort)
ws["D7"].value = ort
"""
from io import BytesIO
import xlsxwriter
workbook=xlsxwriter.Workbook("Image.xlsx")
worksheet=workbook.add_worksheet()
fileName="impro-300.png"
file=open(fileName,"rb")
data=BytesIO(file.read())
file.close()
worksheet.insert_image('A3',fileName,{"image_data":data})
workbook.close() 
