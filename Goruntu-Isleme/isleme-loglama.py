# -*- coding: utf-8 -*-
"""
Created on Sun Nov 20 15:40:34 2022

@author: Ruzgar
"""

import datetime
import cv2 #OPENCV projeye import ediyoruz(Projemize OpenCV Ekliyoruz)
import numpy as np #Görüntü işlemede matematiksel işlermler dolaylı yoldan geldiği için numpy kütüphanesini projemize ekledik.
from matplotlib import  pyplot as plt
from openpyxl import Workbook, load_workbook

try:
    wb =load_workbook("Log.xlsx")
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active
    ws.append(["Application","Information","Value","Application Date"])


image_path="yol.jpeg"
img=cv2.imread(image_path)

def colorFilter(image, value):
    image[:,:,0] = value
    
    log=["ColorFilter", "Color Filter Set To Value : ", value, tarih]
    ws.append(log)
    return image

def YaklasVeCevir(image,yaklas,cevir):
    (h,w)=image.shape[:2]
    center=(w/2,h/2)
    M=cv2.getRotationMatrix2D(center,cevir,scale=yaklas)
    rotated=cv2.warpAffine(img,M,(w,h))
    
    if (yaklas != 1):
        log=["Yaklaş","Scale Changed to : ", yaklas, tarih]
        ws.append(log)
    log=["Çevir","Rotation value : ", cevir,tarih]
    ws.append(log)
    return rotated
    
def makeGray(image):
    Gray=cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    log=["MakeGray", "İmage set to gray filter ", "--", tarih]
    ws.append(log)
    return Gray

def GBlur(image):
    GBlur =cv2.GaussianBlur(image,(5,5),3)
    
    log=["Gaussian Blur", "Gaussian blur applied ", "--", tarih]
    ws.append(log)
    return GBlur

def makeBlur(image,dikey,yatay):
    Blur=cv2.blur(image,(yatay,dikey))
    
    log=["Blur", "Vertical Blur value set to :", yatay, tarih]
    ws.append(log)
    log=["Blur", "Horizontal Blur value set to :", yatay, tarih]
    ws.append(log)
    return Blur

def addLines(image):
    Lines = cv2.HoughLinesP(image, 1, np.pi/180,30)
    for line in Lines:
        x1,y1,x2,y2 = line[0]
        cv2.line(img,(x1,y1),(x2,y2),(0,255,0),4)
    log=["Add Lines", "Lines applied ", "--", tarih]
    ws.append(log)
        
def aynalama(image,boyut):
    aynalama = cv2.copyMakeBorder(image,boyut,boyut,boyut,boyut,cv2.BORDER_REFLECT)
    
    log=["aynalama", "aynalama yapıldı ", "--", tarih]
    ws.append(log)
    return aynalama


def uzatma(image,boyut):
    uzatma = cv2.copyMakeBorder(image,boyut,boyut,boyut,boyut,cv2.BORDER_REPLICATE)
    
    log=["uzatma", "uzatma yapıldı ", "--", tarih]
    ws.append(log)
    return uzatma


def tekrarlama(image,boyut):
    tekrarlama = cv2.copyMakeBorder(image,boyut,boyut,boyut,boyut,cv2.BORDER_WRAP)
    
    log=["tekrarlama", "tekrarlama yapıldı ", "--", tarih]
    ws.append(log)
    return tekrarlama


def sarma(image,boyut):
    sarma = cv2.copyMakeBorder(image,boyut,boyut,boyut,boyut,cv2.BORDER_CONSTANT, value=(75,150,125))  
    
    log=["sarma", "sarma yapıldı ", "--", tarih]
    ws.append(log)
    return sarma    

tarih = datetime.datetime.now()

a = colorFilter(img, 22)
a = YaklasVeCevir(a, 10, 22)

plt.imshow(a)
plt.show()

wb.save("Log.xlsx")

