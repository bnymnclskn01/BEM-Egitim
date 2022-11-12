# -*- coding: utf-8 -*-
"""
Created on Thu Nov 10 18:23:10 2022

@author: Ruzgar
"""
from sys import exit
from openpyxl import Workbook, load_workbook

#Workbook
Başla = input("(1) Varolan dosyaya ekleme yap \n(2) Yeni dosya oluştur (!!eskisini siler)")
if (Başla == "1"):
    Liste = load_workbook("Liste.xlsx")
    sheets = Liste.sheetnames  
    ws1 = Liste[sheets[0]]
    ws2 = Liste[sheets[0]]
    ws3 = Liste[sheets[0]]

elif (Başla == "2"):
    Liste = Workbook()
    ws1 = Liste.active
    ws1.title = "Alt Şirket 1"
    ws2 = Liste.create_sheet(title = 'Alt Şirket 2')
    ws3 = Liste.create_sheet(title = 'Alt şirket 3')

    Başlıklar = ["İd", "İsim", "Yaş","İşe Başlama Tarihi","Departman","Görev"]
    ws1.append(Başlıklar)
    ws2.append(Başlıklar)
    ws3.append(Başlıklar)
else:
    print("Yanlış işlem. Programı tekrar başlatın")
    exit()
    
""" OLMADI
YeniÇalışan = []
YeniAlan = []

class Şirket:
    def GenelBilgi(self,İd,İsim,Yaş,Tarih):
        YeniÇalışan = [İd,İsim,Yaş,Tarih]

class Çalışan(Şirket):
    def AltŞirket1(self,Departman,Görev):
        YeniAlan = [Departman,Görev]

 """      
ÇalışanListesi1 = [] 
ÇalışanListesi2 = [] 
ÇalışanListesi3 = [] 

def ÇalışanEkle():
        YeniÇalışan = [ ]
        YeniÇalışan.insert(1,input("İd : "))
        YeniÇalışan.insert(1,input("İsim : "))
        YeniÇalışan.insert(1,input("Yaş : "))
        YeniÇalışan.insert(1,input("İşe Başlama Tarihi : "))
        YeniÇalışan.insert(1,input("Departman : "))
        YeniÇalışan.insert(1,input("Görev : "))
        return YeniÇalışan
        
        
for i in range(10000000):
    inp = input("(1) Yeni çalışan ekle \n(2) Eklemeyi Bitir \n")
    if(inp == "1"):
        inp2 = input("Alt Şirket Seç \n(1) Alt Şirket 1 \n(2) Alt Şirket 2 \n(3) Alt Şirket 3 \n")
        if(inp2 == "1"):
            ÇalışanListesi1.append(ÇalışanEkle())
        elif(inp2 == "2"):
            ÇalışanListesi2.append(ÇalışanEkle())
        elif(inp2 == "3"):
            ÇalışanListesi3.append(ÇalışanEkle())
        else:
            print("*******Geçersiz İşlem*******")
            
    elif(inp == "2"):
        print("Ekleme Tamamlandı")
        break
    else:
        print("*******Geçersiz İşlem*******")

for row in ÇalışanListesi1:
    ws1.append(row)
for row in ÇalışanListesi2:
    ws2.append(row)
for row in ÇalışanListesi3:
    ws3.append(row)
 

Liste.save("Liste.xlsx")
print("Liste İşlendi") 
